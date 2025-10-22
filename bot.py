"""
Медицинский Telegram-бот: ввод ФИО, опрос по системам, оценка состояния,
рекомендация врача, экспорт отчётов в Excel с цветами, уведомление админа,
возможность повторного опроса.
"""

import logging
import os
from datetime import datetime
import pandas as pd
from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    ConversationHandler,
    ContextTypes,
    filters,
)
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# --------------------------
# ⚙️ Конфигурация
# --------------------------
BOT_TOKEN = os.getenv("TOKEN")
ADMIN_ID = 7146954022
REPORTS_FILENAME = "reports.xlsx"

# --------------------------
# 🧾 Логирование
# --------------------------
logging.basicConfig(
    format="%(asctime)s - %(levelname)s - %(message)s",
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# --------------------------
# Состояния диалога
# --------------------------
ASK_NAME, SELECT_DEPT, ASK_QUESTION = range(3)

# --------------------------
# Вопросы по системам
# --------------------------
DEPARTMENTS = {
    "🫁 Дыхательная система": [
        "Есть ли у вас длительный кашель (больше 2–3 недель)?",
        "Кашель сопровождается мокротой? Какого она цвета?",
        "Чувствуете ли одышку при небольшой нагрузке?",
        "Есть ли боль в груди при дыхании?",
        "Повышается ли температура тела?",
        "Бывают ли ночные приступы кашля или удушья?",
        "Есть ли хрипы при дыхании?",
        "Курите ли вы или контактируете с дымом?",
        "Появляется ли слабость, потливость, снижение массы тела?",
        "Были ли у вас заболевания лёгких ранее?",
    ],
    "🍽 Пищеварительная (ЖКТ)": [
        "Есть ли боли в животе и где именно?",
        "Есть ли тошнота, рвота или изжога?",
        "Менялся ли характер или частота стула?",
        "Есть ли вздутие, урчание, чувство тяжести?",
        "Замечали ли кровь или слизь в стуле?",
        "Есть ли потеря веса или аппетита?",
        "Употребляете ли часто жареное, острое или алкоголь?",
        "Были ли ранее гастрит, язва, колит?",
        "Наблюдается ли желтушность кожи или глаз?",
        "Есть ли горечь во рту или боль под рёбрами?",
    ],
    "💓 Сердечно-сосудистая": [
        "Есть ли боли или жжение за грудиной?",
        "Возникает ли боль при физической нагрузке?",
        "Есть ли одышка при ходьбе или покое?",
        "Бывает ли учащённое сердцебиение?",
        "Бывают ли перебои в сердце?",
        "Есть ли отёки ног по вечерам?",
        "Часто ли повышается давление?",
        "Были ли головокружения или обмороки?",
        "Есть ли хронические заболевания сердца?",
        "Принимаете ли препараты для давления или сердца?",
    ],
    "🧠 Нервная система": [
        "Беспокоят ли головные боли? Где они локализуются?",
        "Есть ли онемение в конечностях или лице?",
        "Бывают ли судороги, дрожь или тремор?",
        "Есть ли нарушения сна или повышенная сонливость?",
        "Бывают ли резкие перепады настроения или тревожность?",
        "Часто ли кружится голова?",
        "Были ли травмы головы?",
        "Есть ли проблемы с памятью, вниманием, речью?",
        "Нарушается ли координация движений?",
        "Бывает ли потеря сознания?",
    ],
    "⚖️ Эндокринная": [
        "Быстрая утомляемость или слабость?",
        "Изменение веса без причины?",
        "Повышенная жажда или частое мочеиспускание?",
        "Изменение настроения, раздражительность?",
        "Часто ли чувствуете холод или жар?",
        "Есть ли сухость кожи, ломкость ногтей, выпадение волос?",
        "Нарушение сна, аппетита или менструального цикла?",
        "Были ли у вас болезни щитовидной железы или диабет?",
        "Есть ли отёки лица или конечностей?",
        "Появляется ли тремор или сердцебиение?",
    ],
    "🦴 Опорно-двигательная": [
        "Есть ли боли в спине, шее или суставах?",
        "Скованность по утрам или после покоя?",
        "Отёки, покраснение или деформация суставов?",
        "Боль усиливается при движении?",
        "Есть ли мышечная слабость?",
        "Проблемы с походкой или равновесием?",
        "Хруст или щелчки при движении?",
        "Были ли травмы костей или связок?",
        "Есть ли утомляемость после физической нагрузки?",
        "Ограничены ли движения в определённых частях тела?",
    ],
}

# --------------------------
# Варианты ответов
# --------------------------
ANSWER_OPTIONS = ["Нет", "Иногда", "Да", "Часто"]
ANSWER_SCORES = {"Нет": 0, "Иногда": 1, "Да": 2, "Часто": 3}

# --------------------------
# Функции оценки
# --------------------------
def calculate_severity(total, max_total):
    ratio = total / max_total
    if ratio < 0.25:
        return "лёгкое"
    elif ratio < 0.6:
        return "среднее"
    else:
        return "тяжёлое"

def recommend_doctor(department, severity):
    doctors = {
        "🫁 Дыхательная система": "Пульмонолог",
        "🍽 Пищеварительная (ЖКТ)": "Гастроэнтеролог",
        "💓 Сердечно-сосудистая": "Кардиолог",
        "🧠 Нервная система": "Невролог",
        "⚖️ Эндокринная": "Эндокринолог",
        "🦴 Опорно-двигательная": "Ортопед / Ревматолог",
    }
    doc = doctors.get(department, "Терапевт")
    if severity == "тяжёлое":
        return f"⚠️ Срочно обратитесь к врачу: {doc}"
    elif severity == "среднее":
        return f"🩺 Желательна консультация: {doc}"
    return f"✅ Ваше состояние стабильное, наблюдение у {doc} при необходимости."

# --------------------------
# Работа с Excel
# --------------------------
def append_report(report_row):
    df_row = pd.DataFrame([report_row])
    if os.path.exists(REPORTS_FILENAME):
        df_old = pd.read_excel(REPORTS_FILENAME, engine="openpyxl")
        df_new = pd.concat([df_old, df_row], ignore_index=True)
        df_new.to_excel(REPORTS_FILENAME, index=False, engine="openpyxl")
    else:
        df_row.to_excel(REPORTS_FILENAME, index=False, engine="openpyxl")

    wb = load_workbook(REPORTS_FILENAME)
    ws = wb.active
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    severity_fill = {
        "лёгкое": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "среднее": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "тяжёлое": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        severity = row[4].value
        if severity in severity_fill:
            for cell in row:
                cell.fill = severity_fill[severity]

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    wb.save(REPORTS_FILENAME)

# --------------------------
# Диалоговые функции
# --------------------------
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Здравствуйте! 👋\nПожалуйста, введите своё ФИО для начала опроса:",
        reply_markup=ReplyKeyboardRemove()
    )
    return ASK_NAME

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Опрос отменён.", reply_markup=ReplyKeyboardRemove())
    return ConversationHandler.END

async def ask_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name = update.message.text.strip()
    if not name:
        await update.message.reply_text("Введите корректное ФИО.")
        return ASK_NAME
    context.user_data["full_name"] = name
    keyboard = [[d] for d in DEPARTMENTS.keys()] + [["Выход"]]
    await update.message.reply_text(
        f"Спасибо, {name}! Выберите систему организма для опроса:",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
    )
    return SELECT_DEPT

async def select_dept(update: Update, context: ContextTypes.DEFAULT_TYPE):
    dept = update.message.text

    if dept == "Пройти ещё один опрос":
        context.user_data["answers"] = {}
        context.user_data["index"] = 0
        context.user_data["department"] = None
        context.user_data["questions"] = []

        keyboard = [[d] for d in DEPARTMENTS.keys()] + [["Выход"]]
        await update.message.reply_text(
            "Выберите систему организма для нового опроса:",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        )
        return SELECT_DEPT

    if dept == "Выход":
        await update.message.reply_text("Спасибо! До свидания 👋", reply_markup=ReplyKeyboardRemove())
        return ConversationHandler.END

    if dept not in DEPARTMENTS:
        await update.message.reply_text("Пожалуйста, выберите систему из списка.")
        return SELECT_DEPT

    context.user_data["department"] = dept
    context.user_data["questions"] = DEPARTMENTS[dept]
    context.user_data["answers"] = {}
    context.user_data["index"] = 0

    q = context.user_data["questions"][0]
    keyboard = ReplyKeyboardMarkup([[a] for a in ANSWER_OPTIONS] + [["Выход"]], resize_keyboard=True)
    await update.message.reply_text(f"Вопрос 1/10:\n{q}", reply_markup=keyboard)
    return ASK_QUESTION

async def ask_question(update: Update, context: ContextTypes.DEFAULT_TYPE):
    answer = update.message.text
    if answer == "Выход":
        return await cancel(update, context)
    if answer not in ANSWER_OPTIONS:
        await update.message.reply_text("Выберите вариант ответа с кнопок ниже.")
        return ASK_QUESTION

    idx = context.user_data["index"]
    questions = context.user_data["questions"]
    context.user_data["answers"][questions[idx]] = answer
    idx += 1
    context.user_data["index"] = idx

    if idx < len(questions):
        q = questions[idx]
        keyboard = ReplyKeyboardMarkup([[a] for a in ANSWER_OPTIONS] + [["Выход"]], resize_keyboard=True)
        await update.message.reply_text(f"Вопрос {idx+1}/10:\n{q}", reply_markup=keyboard)
        return ASK_QUESTION

    dept = context.user_data["department"]
    answers = context.user_data["answers"]
    total = sum(ANSWER_SCORES[a] for a in answers.values())
    max_total = len(answers) * 3
    severity = calculate_severity(total, max_total)
    doctor = recommend_doctor(dept, severity)

    report = {
        "Дата": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "ID пользователя": update.effective_user.id,
        "ФИО": context.user_data["full_name"],
        "Система": dept,
        "Степень": severity,
        "Рекомендовано": doctor,
        "Баллы": total,
        "Из": max_total,
    }
    append_report(report)

    summary = (
        f"📋 *Опрос завершён*\n\n"
        f"ФИО: {context.user_data['full_name']}\n"
        f"Система: {dept}\n"
        f"Состояние: *{severity}*\n\n"
        f"{doctor}"
    )
    keyboard = [["Пройти ещё один опрос"], ["Выход"]]
    await update.message.reply_text(
        summary,
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
        parse_mode="Markdown"
    )

    admin_summary = (
        f"🧾 Новый отчёт:\n\n"
        f"ФИО: {context.user_data['full_name']}\n"
        f"Система: {dept}\n"
        f"Сумма баллов: {total}/{max_total}\n"
        f"Состояние: {severity}\n"
        f"{doctor}"
    )
    await context.bot.send_message(chat_id=ADMIN_ID, text=admin_summary)

    if os.path.exists(REPORTS_FILENAME):
        with open(REPORTS_FILENAME, "rb") as f:
            await context.bot.send_document(chat_id=ADMIN_ID, document=f, filename=REPORTS_FILENAME)

    return SELECT_DEPT

async def export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ У вас нет доступа к этой команде.")
        return
    if not os.path.exists(REPORTS_FILENAME):
        await update.message.reply_text("📁 Пока нет данных для экспорта.")
        return
    with open(REPORTS_FILENAME, "rb") as f:
        await update.message.reply_document(f, filename=REPORTS_FILENAME)

# --------------------------
# Запуск бота
# --------------------------
def main():
    app = ApplicationBuilder().token(BOT_TOKEN).build()

    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            ASK_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_name)],
            SELECT_DEPT: [MessageHandler(filters.TEXT & ~filters.COMMAND, select_dept)],
            ASK_QUESTION: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_question)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    app.add_handler(conv_handler)
    app.add_handler(CommandHandler("export", export))

    logger.info("🤖 Бот запущен и готов к работе.")
    app.run_polling()

if __name__ == "__main__":
    main()
